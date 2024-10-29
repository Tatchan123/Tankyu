import jikken.gpu as gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from collections import OrderedDict
from data.load import load_mnist
import openpyxl
import copy
from model1 import *
from trainer import *

"""ここはエクセル打ち込みとか発表するための実験をするとこってかんじで"""

"""
    self, step, layer, weightinit, optimizer, data, batch_size, lr, check, (decreace1, decreace2, epsilon, complement, rmw_layer, delete_n)
"""
print("start")

data_n = 1000
(x_train, t_train),(x_test,t_test) = load_mnist(normalize=True)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:data_n], "t_test":t_test[:data_n]}

sgd_result = []
onlytoba_result = []
complement_result = []
randomrmw_result = []
for i in range(100):
    print(str(i),"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
    sgd = Trainer(step=[200], layer=[100,100,100], weightinit=He, optimizer="sgd", data=data, batch_size=100, lr=0.04, check=50)
    sgd_result.append(sgd.fit())

    parameter = sgd.params

    otparams = copy.deepcopy(parameter)
    onlytoba = Trainer(step=["rmw"], layer=[100,100,100], weightinit=otparams, optimizer="sgd", data=data, batch_size=100, lr=0.001, 
                     check=50, epsilon=[1e-6,1e-2,1e-2,1.2e-2], complement=False, rmw_layer=[2,3,4])
    onlytoba_result.append(onlytoba.fit())

    cpparams = copy.deepcopy(parameter)
    complement = Trainer(step=["rmw"], layer=[100,100,100], weightinit=cpparams, optimizer="sgd", data=data, batch_size=100, lr=0.001, 
                     check=50, epsilon=[1e-6,1e-2,1e-2,1.2e-2], complement=True, rmw_layer=[2,3,4])
    complement_result.append(complement.fit())

    rrparams = copy.deepcopy(parameter)
    randomrmw = Trainer(step=["random_rmw"], layer=[100,100,100], weightinit=rrparams, optimizer="sgd", data=data, batch_size=100, lr=0.001, 
                     check=50, epsilon=[1e-6,1e-2,1e-2,1.2e-2], complement=False, rmw_layer=[2,3,4],delete_n=[0,25,20,15])
    randomrmw_result.append(randomrmw.fit())

print(sgd_result)
print(onlytoba_result)
print(complement_result)
print(randomrmw_result)

wb = openpyxl.load_workbook('100result.xlsx')
sheet = wb['Sheet1']
for i in range(len(sgd_result)):
    sheet.cell(row=2, column=i+3).value = sgd_result[i][0]
    sheet.cell(row=6, column=i+3).value = sgd_result[i][1]
for i in range(len(onlytoba_result)):
    sheet.cell(row=3, column=i+3).value = onlytoba_result[i][0]
    sheet.cell(row=7, column=i+3).value = onlytoba_result[i][1]
for i in range(len(complement_result)):
    sheet.cell(row=4, column=i+3).value = complement_result[i][0]
    sheet.cell(row=8, column=i+3).value = complement_result[i][1]
for i in range(len(randomrmw_result)):
    sheet.cell(row=5, column=i+3).value = randomrmw_result[i][0]
    sheet.cell(row=9, column=i+3).value = randomrmw_result[i][1]

sheet.cell(row=1, column=1).value = 0
wb.save(r'100result2.xlsx')
wb.close()