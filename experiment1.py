import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from data.load import load_mnist
from load_cifar10 import load_cifar10

from model import *
from trainer import *
from params_init import *
import pickle
import openpyxl as pyxl

data_n = 512


#x_train, t_train, x_test, t_test = load_cifar10(normalize=True, means=[0.5,0.5,0.5], stds=[0.5,0.5,0.5])
(x_train, t_train),(x_test, t_test) = load_mnist(normalize=True,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:512], "t_test":t_test[:512]}


layer1 = [512,256,128]
conv_layer1 = [[32,3,1],[2],[64,3,1],[2],[128,3,1],[2]]
opt2 = {"opt":"adam", "dec1":0.9, "dec2":0.999,"lr":0.002,"batch_size":64}
exp = {"method":"exp", "base":0.92}
#toba_option = {"rmw_type":"coco_toba","epsilon":[0.0,0.0,0.0],"complement":True,"rmw_layer":["Affine1","Affine2","Affine3"],"delete_n":[156,78,39],"strict":True}


random_results = []
coco_results = []
fit_results = []

while True:
    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=Relu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    base = Trainer(network, optimizer=opt2, data=data, check=5, scheduler=exp)
    base.fit(10)
    base.coco_sort(["Affine2","Affine3","Affine4"])
    for delper in [0.1,0.5]:
        dels = [int(2048*delper),int(512*delper),int(256*delper),int(128*delper)]
        
        random = copy.deepcopy(base).rmw_fit("random_rmw",["Affine2","Affine3","Affine4"],dels)
        random_results.append(random)
        
        cocotest = copy.deepcopy(base)
        coco = cocotest.rmw_fit("coco_toba",["Affine2","Affine3","Affine4"],dels,[0.0,0.0,0.0,0.0])
        coco_results.append(coco)
        
        fit_results.append(cocotest.fit(5))
        
    with open('exp1result.pkl','wb') as f:
        pickle.dump(random_results, f)
        pickle.dump(coco_results, f)
        pickle.dump(fit_results, f)


for delper in [0.1]:
    dels = [int(2048*delper),int(512*delper),int(256*delper),int(128*delper)]
    

    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=Relu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    base = Trainer(network, optimizer=opt2, data=data, check=5, scheduler=exp)
    base.fit(10)


    random = copy.deepcopy(base).rmw_fit("random_rmw",["Affine1","Affine2","Affine3","Affine4"],dels)
    random_results.append(random)

    coco_option = {"rmw_type":"coco_toba","epsilon":[0.0,0.0,0.0,0.0],"complement":True,"rmw_layer":["Affine1","Affine2","Affine3","Affine4"],"delete_n":dels,"strict":True}
    coco_model = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=copy.deepcopy(global_params), activation=Relu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    coco_toba = Trainer(coco_model, step=['toba'], optimizer=opt2, data=data, check=5, tobaoption=coco_option,scheduler=exp)
    coco_dacc, coco_acc2 = coco_toba.fit()
    coco_results.append([coco_dacc,coco_acc2])


    fit_toba = Trainer(coco_model, step=[5], optimizer=opt2, data=data, check=5, tobaoption=coco_option,scheduler=exp)
    fit_dacc,cocofit_acc2 = fit_toba.fit()
    fit_results.append([fit_dacc,cocofit_acc2])

final_result = {"random_result":random_results,"coco_result":coco_results,"cocofit_result":fit_results}

with open('exp1result.pkl','wb') as f:
    pickle.dump(final_result, f)


wb = pyxl.load_workbook('result1.xlsx')
sheet = wb['Sheet1']

for i in range(random_results):
    sheet.cell(2,i+1).value = random_results[i][0]
    sheet.cell(3,i+1).value = random_results[i][1]

    sheet.cell(4,i+1).value = coco_results[i][0]
    sheet.cell(5,i+1).value = coco_results[i][1]

    sheet.cell(6,i+1).value = fit_results[i][0]
    sheet.cell(7,i+1).value = fit_results[i][1]

wb.save()
wb.close()
