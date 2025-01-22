import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from data.load import load_mnist
from load_cifar10 import load_cifar10

from model import *
from trainer import *
from params_init import *
import pickle
import openpyxl as pyxl
import threading


data_n = 16384


x_train, t_train, x_test, t_test = load_cifar10(normalize=True, means=[0.5,0.5,0.5], stds=[0.5,0.5,0.5])
#(x_train, t_train),(x_test, t_test) = load_mnist(normalize=True,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:512], "t_test":t_test[:512]}


layer1 = [512,256,128]
conv_layer1 = [[32,3,1],[2],[64,3,1],[2],[128,3,1],[2]]
opt2 = {"opt":"adam", "dec1":0.9, "dec2":0.999,"lr":0.002,"batch_size":64}
exp = {"method":"exp", "base":0.92}

random_results = []
coco_results = []
fit_results = []

random_tmp = []
coco_tmp = []
fit_tmp = []

def input_quit():
    while True:
        user_input = input("type 'q' to quit")
        if user_input.lower() == 'q':
            global running
            running = False
            break


running = True
input_quit = threading.Thread(target=input_quit)
input_quit.start()



while running:
    i = 0
    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=Relu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    base = Trainer(network, optimizer=opt2, data=data, check=5, scheduler=exp)
    base.fit(10)
    base.coco_sort(["Affine2","Affine3","Affine4"])
    for delper in [0.1,0.3,0.5,]:
        dels = [int(2048*delper),int(512*delper),int(256*delper),int(128*delper)]

        
        random = copy.deepcopy(base).rmw_fit("random_rmw",["Affine2","Affine3","Affine4"],dels)
        random_tmp.append(random["acc"])
        
        cocotest = copy.deepcopy(base)
        coco = cocotest.rmw_fit("coco_toba",["Affine2","Affine3","Affine4"],dels,[0.0,0.0,0.0,0.0])
        coco_tmp.append(coco["acc"])
        
        fit_tmp.append(cocotest.fit(5))
    
    random_results.append(random_tmp)
    coco_results.append(coco_tmp)
    fit_results.append(fit_tmp)
    
    with open('exp1result.pkl','wb') as f:
        pickle.dump(random_results, f)
        pickle.dump(coco_results, f)
        pickle.dump(fit_results, f)
    


    wb = pyxl.load_workbook('result1.xlsx')
    sheet = wb['Sheet1']
    for j in range(len(random_results[0])):
        sheet.cell(row=j+2, column=i+4).value = random_results[-1][j]

        sheet.cell(row=j+2+len(random_results[0]), column=i+4).value = coco_results[-1][j]

        sheet.cell(row=j+2+2*len(random_results[0]), column=i+4).value = fit_results[-1][j]


    wb.save('result1.xlsx')
    wb.close()
    print("saved")
    i += 1
