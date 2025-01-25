import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from data.load import load_mnist
from load_cifar10 import load_cifar10

from model import *
from trainer import *
from params_init import *
import openpyxl as pyxl
import threading

data_n = 8192


#x_train, t_train, x_test, t_test = load_cifar10(normalize=True, means=[0.5,0.5,0.5], stds=[0.5,0.5,0.5])
(x_train, t_train),(x_test, t_test) = load_mnist(normalize=True,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:256], "t_test":t_test[:256]}


layer1 = [512,256,128]
conv_layer1 = [[32,3,1],[2],[64,3,1],[2],[128,3,1],[2]]
opt2 = {"opt":"adam", "dec1":0.9, "dec2":0.999,"lr":0.002,"batch_size":64}
exp = {"method":"exp", "base":0.92}



def input_quit():
    while True:
        user_input = input("type 'q' to quit\n")
        if user_input.lower() == 'q':
            global running
            running = False
            break


running = True
input_quit = threading.Thread(target=input_quit)
input_quit.start()



i = 4


while running:
    result = []
    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    test = Trainer(network, optimizer=opt2, data=data, check=1, scheduler=exp)
    test.fit(10)

    for eps in [0.75,0.7,0.65,0.6,0.5]:
        cp = copy.deepcopy(test)
        print("epsilon =",eps)
        cp.epsilon_coco_sort([eps,eps,eps,eps],["Affine1","Affine2","Affine3","Affine4"])
        accloss = cp.rmw_fit("coco_toba",["Affine1","Affine2","Affine3","Affine4"],[2048,512,256,128])
        result.append(accloss)


    wb = pyxl.load_workbook('epsilon.xlsx')
    sh = wb['Sheet1']

    for j in range(len(result)):
        sh.cell(row=j*3+2, column=i+4).value = result[j][0]
        sh.cell(row=j*3+3, column=i+4).value = result[j][1]
    wb.save('epsilon.xlsx')
    wb.close()

    i += 1
