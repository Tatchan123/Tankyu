import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from collections import OrderedDict
from data.load import load_mnist
import openpyxl
import copy
from model1 import *
from trainer import *

"""
    self, step, layer, weightinit, optimizer, data, batch_size, lr, check, (decreace1, decreace2, epsilon, complement, rmw_layer, delete_n)
"""
print("start")



data_n = 2000
(x_train, t_train),(x_test,t_test) = load_mnist(normalize=True)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:data_n], "t_test":t_test[:data_n]}

result = {}
layer_size = [15,30,50,100]
delete_list = [10,15,20,25,30,40,50]
t = 10



for layer_n in [1,2,3]:
    for layer_size in [15,30,50,100]:
        layer = np.full(layer_n,layer_size)
        layer = layer.tolist()
        for n in range(t):
            d0 = Trainer(step=[200], layer=layer, weightinit=He, optimizer="sgd", data=data, batch_size=100, lr=0.04, check=50)
            if str(layer_size)+"_"+str(layer_n)+"_"+str(0) in result:
                result[str(layer_size)+"_"+str(layer_n)+"_"+str(0)].append(d0.fit())
            else:    result[str(layer_size)+"_"+str(layer_n)+"_"+str(0)] = [d0.fit()]
            params = d0.params
            rmwlayer = list(range(2,2+layer_n))
            for i in range(len(delete_list)):
                rmw_n = round(layer_size*delete_list[i]/100)
                pa = copy.deepcopy(params)
                dn = Trainer(step=["count_rmw"], layer=layer, weightinit=pa, optimizer="sgd", data=data, batch_size=100, lr=0.04, 
                            check=50,epsilon=[1e-6,1e-2,1e-2,1.2e-2], complement=True, rmw_layer=rmwlayer,rmw_n=rmw_n)
                if str(layer_size)+"_"+str(layer_n)+"_"+str(i) in result:
                    result[str(layer_size)+"_"+str(layer_n)+"_"+str(i)].append(dn.fit())
                else:   result[str(layer_size)+"_"+str(layer_n)+"_"+str(i)] = [dn.fit()]
    print(layer_n,layer_size)
wb = openpyxl.load_workbook('100result.xlsx')
sheet = wb['Sheet1']
r = 6
c = 6
for i in range(8):
    for j in [1,2,3]:
        for k in [15,30,50,100]:
            sheet.cell(row=r, column=c).value = str(i)+str(j)+str(k)
            for result in result[str(k)+"_"+str(j)+"_"+str(i)]:
                c += 1
                sheet.cell(row=r, column=c).value = result[0]
            r += 1
            c = 6
            
sheet.cell(row=6, column=i+3).value



wb.save('delete_test.xlsx')
wb.close()
            