import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from collections import OrderedDict
from data.load import load_mnist
import openpyxl
import copy
from model1 import *
from trainer import *

"""ここは発表の「実験１」用で"""

"""
    self, step, layer, weightinit, optimizer, data, batch_size, lr, check, (decreace1, decreace2, epsilon, complement, rmw_layer, delete_n)
"""
print("start")

data_n = 6400
(x_train, t_train),(x_test,t_test) = load_mnist(normalize=False,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:data_n], "t_test":t_test[:data_n]}
for i in range(100):
    if t_train[i] == 4:
        print(x_train[i])
        x = x_train[i].reshape(28,28)
        wb = openpyxl.load_workbook('sss.xlsx')
        sheet = wb["Sheet1"]
        for i in range(28):
            for j in range(28):
                sheet.cell(row=i, column=j).value = x[i][j]
        wb.close()
        break
t = np.argmax(data["t_test"],axis=1)


sgd_result = []

toba_result = []
randomrmw_result = []
for i in range(100):
    print(str(i),"aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
    sgd = Trainer(step=[200], layer=[100,100,100], weightinit=He, optimizer="sgd", data=data, batch_size=64, lr=0.04, check=50)
    sgd_result.append(sgd.fit())

    parameter = sgd.params

    cpparams = copy.deepcopy(parameter)
    complement = Trainer(step=["rmw"], layer=[100,100,100], weightinit=cpparams, optimizer="sgd", data=data, batch_size=64, lr=0.001, 
                     check=50, epsilon=[1e-6,3e-3,1.5e-3,1.2e-3], complement=True, rmw_layer=[2,3,4])
    toba_result.append(complement.fit())

    rrparams = copy.deepcopy(parameter)
    randomrmw = Trainer(step=["random_rmw"], layer=[100,100,100], weightinit=rrparams, optimizer="sgd", data=data, batch_size=64, lr=0.001, 
                     check=50, rmw_layer=[2,3,4],delete_n=[0,10,10,7])
    randomrmw_result.append(randomrmw.fit())

print(sgd_result)
print(toba_result)
print(randomrmw_result)

wb = openpyxl.load_workbook('100result1.xlsx')
sheet = wb['Sheet1']
for i in range(len(sgd_result)):
    sheet.cell(row=2, column=i+3).value = sgd_result[i][0]
    sheet.cell(row=5, column=i+3).value = sgd_result[i][1]
for i in range(len(toba_result)):
    sheet.cell(row=3, column=i+3).value = toba_result[i][0]
    sheet.cell(row=6, column=i+3).value = toba_result[i][1]
for i in range(len(randomrmw_result)):
    sheet.cell(row=4, column=i+3).value = randomrmw_result[i][0]
    sheet.cell(row=7, column=i+3).value = randomrmw_result[i][1]

sheet.cell(row=1, column=1).value = 0
wb.save('100result1.xlsx')
wb.close()