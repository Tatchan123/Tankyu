import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from data.load import load_mnist
from load_cifar10 import load_cifar10

from model import *
from trainer import *
from params_init import *
import pickle
import openpyxl as pyxl
import threading


data_n = 32768


x_train, t_train, x_test, t_test = load_cifar10(normalize=True, means=[0.5,0.5,0.5], stds=[0.5,0.5,0.5])
#(x_train, t_train),(x_test, t_test) = load_mnist(normalize=True,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:256], "t_test":t_test[:256]}


layer1 = [512,256,128]
conv_layer1 = [[32,3,1],[2],[64,3,1],[2],[128,3,1],[2]]
opt2 = {"opt":"adam", "dec1":0.9, "dec2":0.999,"lr":0.002,"batch_size":64}
exp = {"method":"exp", "base":0.92}

layer_names = ["Affine1","Affine2","Affine3","Affine4"]

i = 0

while True:
    random_result, nozero_random_result, coco_result, zero_coco_result = [],[],[],[]


    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=Relu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    model = test = Trainer(network, optimizer=opt2, data=data, check=5, scheduler=exp)
    model.fit(50)
    model.coco_sort(layer_names)
    for delper in [0.0,0.1,0.2,0.3,0.4,0.5]:
        delete_ns = [int(2048*delper),int(512*delper),int(256*delper),int(128*delper)]

        rt = copy.deepcopy(model).rmw_fit("random_toba",layer_names,delete_ns)[0]
        random_result.append(rt)

        nzr = copy.deepcopy(model).rmw_fit("nozero_toba",layer_names,delete_ns)[0]
        nozero_random_result.append(nzr)

        ct = copy.deepcopy(model).rmw_fit("coco_toba",layer_names,delete_ns)[0]
        coco_result.append(ct)

        zct = copy.deepcopy(model).rmw_fit("zero_coco",layer_names,delete_ns)[0]
        zero_coco_result.append(zct)
    
    wb = pyxl.load_workbook("zero-cifar10.xlsx")
    st = wb['Sheet1']

    for j in range(len(random_result)):
        st.cell(row=j+2, column=i+4).value = random_result[j]
        st.cell(row=j+4+len(random_result), column=i+4).value = nozero_random_result[j]
        st.cell(row=j+6+2*len(random_result), column=i+4).value = coco_result[j]
        st.cell(row=j+8+3*len(random_result), column=i+4).value = zero_coco_result[j]

    wb.save('zero-cifar10.xlsx')
    wb.close()
    i += 1

