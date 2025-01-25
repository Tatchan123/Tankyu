import gpu
if gpu.Use_Gpu:
    import cupy as np
else:
    import numpy as np
    
from data.load import load_mnist
from load_cifar10 import load_cifar10

from model import *
from trainer import *
from params_init import *
import threading
import openpyxl as pyxl


data_n = 8192


#x_train, t_train, x_test, t_test = load_cifar10(normalize=True, means=[0.5,0.5,0.5], stds=[0.5,0.5,0.5])
(x_train, t_train),(x_test, t_test) = load_mnist(normalize=True,flatten=False)
data = {"x_train":x_train[:data_n], "t_train":t_train[:data_n], "x_test":x_test[:256], "t_test":t_test[:256]}


layer1 = [512,256,128]
conv_layer1 = [[32,3,1],[2],[64,3,1],[2],[128,3,1],[2]]
opt2 = {"opt":"adam", "dec1":0.9, "dec2":0.999,"lr":0.002,"batch_size":64}
exp = {"method":"exp", "base":0.92}

def input_quit():
    while True:
        user_input = input("type 'q' to quit\n")
        if user_input.lower() == 'q':
            global running
            running = False
            break


running = True
input_quit = threading.Thread(target=input_quit)
input_quit.start()


i = 0

while running:
    print("DropReg")
    DropReg_result = []
    No = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True)
    No_train = Trainer(No, optimizer=opt2, data=data, check=10, scheduler=exp)
    No_train.fit(10)
    DropReg_result.append(No_train.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))

    Drop = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True,drop_rate=[0.26,0.33])
    Drop_train = Trainer(Drop, optimizer=opt2, data=data, check=10, scheduler=exp)
    Drop_train.fit(10)
    DropReg_result.append(Drop_train.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))
    

    Reg = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True,regularize=["l2",0.0005])
    Reg_train = Trainer(Reg, optimizer=opt2, data=data, check=10, scheduler=exp)
    Reg_train.fit(10)
    DropReg_result.append(Reg_train.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))

    DropReg = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    DropReg_train = Trainer(DropReg, optimizer=opt2, data=data, check=10, scheduler=exp)
    DropReg_train.fit(10)
    DropReg_result.append(DropReg_train.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))

    wb = pyxl.load_workbook('top10.xlsx')

    DropReg_sheet = wb['DropReg']

    for j in range(4):
        for k in range(10):
            DropReg_sheet.cell(j*5+2,k+i*11+4).value = DropReg_result[j]["Affine1"][k]
            DropReg_sheet.cell(j*5+3,k+i*11+4).value = DropReg_result[j]["Affine2"][k]
            DropReg_sheet.cell(j*5+4,k+i*11+4).value = DropReg_result[j]["Affine3"][k]
            DropReg_sheet.cell(j*5+5,k+i*11+4).value = DropReg_result[j]["Affine4"][k]

    wb.save('top10.xlsx')
    wb.close()
    DropReg_result = []

    print("Node")
    node_result = []
    for mrg in [0.25,0.5,1,2,4]:
        network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=[int(n*mrg) for n in layer1], conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
        test = Trainer(network, optimizer=opt2, data=data, check=10, scheduler=exp)
        test.fit(10)
        node_result.append(test.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))

    wb = pyxl.load_workbook('top10.xlsx')
    node_sheet = wb['Node']

    for j in range(5):
        for k in range(10):
            node_sheet.cell(j*6+2,k+i*11+4).value = node_result[j]["Affine1"][k]
            node_sheet.cell(j*6+3,k+i*11+4).value = node_result[j]["Affine2"][k]
            node_sheet.cell(j*6+4,k+i*11+4).value = node_result[j]["Affine3"][k]
            node_sheet.cell(j*6+5,k+i*11+4).value = node_result[j]["Affine4"][k]
        
    wb.save('top10.xlsx')
    wb.close()
    node_result = []
    print("Optimizer")
    opt_result = []
    opt1 = {"opt":"sgd","lr":0.02,"batch_size":64}

    network = Convnetwork(input_size=(list(x_train[0].shape)), output_size=10, dense_layer=layer1, conv_layer=conv_layer1, weightinit=He, activation=LeakyRelu, batchnorm=True, toba=True, drop_rate=[0.26,0.33], regularize=["l2",0.0005])
    sgd = Trainer(network, optimizer=opt1, data=data, check=10, scheduler=exp)
    sgd.fit(10)
    opt_result.append(sgd.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))
    adam = Trainer(network, optimizer=opt2, data=data, check=10, scheduler=exp)
    adam.fit(10)
    opt_result.append(adam.coco_sort([10,10,10,10],["Affine1","Affine2","Affine3","Affine4"]))

    wb = pyxl.load_workbook('top10.xlsx')
    opt_sheet = wb['Optimizer']
    for j in range(2):
        for k in range(10):
            opt_sheet.cell(j*3+2,k+i*11+4).value = opt_result[j]["Affine1"][k]
            opt_sheet.cell(j*3+3,k+i*11+4).value = opt_result[j]["Affine2"][k]
            opt_sheet.cell(j*3+4,k+i*11+4).value = opt_result[j]["Affine3"][k]
            opt_sheet.cell(j*3+5,k+i*11+4).value = opt_result[j]["Affine4"][k]
    
    i += 1
    opt_result = []